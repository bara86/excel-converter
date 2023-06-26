from __future__ import annotations

import csv
import os
import tempfile
from pathlib import Path
from typing import TYPE_CHECKING, NamedTuple
from enum import IntEnum, auto

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from PySide6.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFileDialog, \
    QMessageBox

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


class Colonne(IntEnum):
    ID = 1
    DATA = auto()
    DATA_INVIO = auto()
    PROFESSIONE = auto()
    SOC = auto()
    SOS = auto()
    ZONA = auto()
    TIPOLOGIA_PRESIDIO = auto()
    SETTING = auto()
    SEDE = auto()
    REQUISITO = auto()
    INDICATORE = auto()
    NUMERATORE = auto()
    DENOMINATORE = auto()
    PERCENTUALE = auto()
    PESO = auto()
    PERCENTUALE_PESATA = auto()


class Indicatore(NamedTuple):
    stringa_uscita: str
    nome_indicatore: str
    peso: int


mappa_requisito_indicatori: dict[str, list[Indicatore]] = {
    'IDENTIFICAZIONE UTENTE': [
        Indicatore(stringa_uscita='IDENTIFICAZIONE ATTIVA',
                   nome_indicatore='ident_utente', peso=2)],
    'PREVENZIONE CADUTE': [
        Indicatore(stringa_uscita='SICUREZZA AMBIENTIE PRESIDI',
                   nome_indicatore='lista_item', peso=3),
        Indicatore(stringa_uscita='POST CADUTA',
                   nome_indicatore='cadute', peso=3)],
    'SORVEGLIANZA INFEZIONI': [
        Indicatore(stringa_uscita='GEL LAVAMANI POSTAZIONE',
                   nome_indicatore='gel', peso=4),
        Indicatore(stringa_uscita='GEL LAVAMANI BORSE',
                   nome_indicatore='borse', peso=4),
        Indicatore(stringa_uscita='GUANTI POSTAZIONE',
                   nome_indicatore='guanti', peso=4),
        Indicatore(stringa_uscita='GUANTI BORSE',
                   nome_indicatore='guanti_borse', peso=4),
        Indicatore(stringa_uscita='POSTER MANI PULITE',
                   nome_indicatore='poster', peso=4),
        Indicatore(stringa_uscita='POSTER MANI NUDE',
                   nome_indicatore='maninude', peso=4),
        Indicatore(stringa_uscita='AZIONI POST MONITORAGGIO',
                   nome_indicatore='azioni_mani', peso=4)],
    'DISPOSITIVI MEDICI': [
        Indicatore(stringa_uscita='MANUALE IN ITALIANO',
                   nome_indicatore='lingua', peso=4),
        Indicatore(stringa_uscita='PROGRAMMAZIONE MANUTENZIONE ESTERNA',
                   nome_indicatore='piano', peso=4),
        Indicatore(stringa_uscita='EFFETTUAZIONE MANUTENZIONE ESTERNA',
                   nome_indicatore='manutenzione', peso=4),
        Indicatore(stringa_uscita='MANUTENZIONE TSLB',
                   nome_indicatore='tslb', peso=4)],
    'SICUREZZA EMOCOMPONENTI': [
        Indicatore(stringa_uscita="CONFORMITA' RICHIESTE",
                   nome_indicatore='nc', peso=3),
        Indicatore(stringa_uscita='ETICHETTATURA CAMPIONE STOCCATO',
                   nome_indicatore='stoc', peso=3)],
    'SICUREZZA PZ ONCOLOGICO': [
        Indicatore(stringa_uscita='PRESCRIZIONE FARMACI CTA',
                   nome_indicatore='presc_CTA', peso=5),
        Indicatore(stringa_uscita='PREPARAZIONE CTA',
                   nome_indicatore='prep_CTA', peso=5)],
    'RISCHIO FARMACI': [
        Indicatore(stringa_uscita='ETICHETTATURA LASA',
                   nome_indicatore='nc_LASA', peso=5),
        Indicatore(stringa_uscita='ALLOCAZIONE LASA',
                   nome_indicatore='stoc_LASA', peso=5)],
    "CONTROLLO QUALITA'": [
        Indicatore(stringa_uscita='CQ APPARECCHIATURE',
                   nome_indicatore='checklist', peso=10)],
    'NEOASSUNTO/NEOINSERITO': [
        Indicatore(stringa_uscita='PIANO INSERIMENTO NEOASSUNTO',
                   nome_indicatore='neoassunto', peso=1),
        Indicatore(stringa_uscita='PIANO INSERIMENTO NEOINSERITO',
                   nome_indicatore='neoinserito', peso=1)]
}


def cercaRigaColonna(sheet: Worksheet, nome_colonna: str) -> str | None:
    for col in range(1, sheet.max_column + 1):
        col_name = get_column_letter(col)
        valore = sheet[f"{col_name}1"].value.strip().lower()
        if valore == nome_colonna.lower() or valore == f"{nome_colonna.lower()}_":
            return col_name
    else:
        assert False, f"Impossibile trovare {nome_colonna}"


def scriviTitoliUscita(sheet: Worksheet) -> None:
    for colonna in Colonne:
        sheet[f'{get_column_letter(colonna)}1'].value = colonna.name.replace('_', ' ').replace('PERCENTUALE', '%')


def converti(path_foglio_ingresso: Path, path_foglio_uscita: Path) -> None:

    _path_foglio_ingresso = path_foglio_ingresso
    is_csv_file = path_foglio_ingresso.suffix == '.csv'
    if is_csv_file:
        _path_foglio_ingresso = Path(f"{tempfile.mkstemp()[1]}.xlsx")

        wb = Workbook()
        ws = wb.active

        with open(path_foglio_ingresso, encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)

        wb.save(str(_path_foglio_ingresso))
    wb: Workbook = load_workbook(str(_path_foglio_ingresso))

    sheet: Worksheet = wb.worksheets[0]
    indice_id = cercaRigaColonna(sheet, 'id')
    indice_zona_presidio = cercaRigaColonna(sheet, 'zona_presidio')
    indice_professione = cercaRigaColonna(sheet, 'Professione')
    indice_data = cercaRigaColonna(sheet, 'data')
    indice_data_invio = cercaRigaColonna(sheet, 'created')
    indice_sos = cercaRigaColonna(sheet, 'SOS')
    indice_soc = cercaRigaColonna(sheet, 'SOC')
    indice_tipologia_presidio = cercaRigaColonna(sheet, 'presidio')
    indice_servizio = cercaRigaColonna(sheet, 'servizio')
    indice_sede_presidio = cercaRigaColonna(sheet, 'sede_presidio')

    wb_uscita = Workbook()
    sheet_uscita = wb_uscita.worksheets[0]

    scriviTitoliUscita(sheet_uscita)

    riga_uscita = 2
    for row in range(2, sheet.max_row + 1):

        for (requisito, indicatori) in mappa_requisito_indicatori.items():
            id_riga = sheet[f"{indice_id}{row}"].value
            zona_presidio_riga = sheet[f"{indice_zona_presidio}{row}"].value[len('zona '):]
            data_riga = sheet[f"{indice_data}{row}"].value
            data_invio_riga = sheet[f"{indice_data_invio}{row}"].value
            servizio_riga = sheet[f'{indice_servizio}{row}'].value
            professione_riga = sheet[f'{indice_professione}{row}'].value
            soc_riga = sheet[f'{indice_soc}{row}'].value
            sos_riga = sheet[f'{indice_sos}{row}'].value
            tipologia_presidio_riga = sheet[f'{indice_tipologia_presidio}{row}'].value[len('presidio '):].capitalize()
            sede_presidio_riga = sheet[f'{indice_sede_presidio}{row}'].value

            for contatore_uscita, indicatore in enumerate(indicatori):
                colonna_num = cercaRigaColonna(sheet, f"num_{indicatore.nome_indicatore}")
                colonna_den = cercaRigaColonna(sheet, f'den_{indicatore.nome_indicatore}')
                colonna_percentuale = cercaRigaColonna(sheet, f'%_{indicatore.nome_indicatore}')
                if colonna_num is None or colonna_den is None or colonna_percentuale is None:
                    continue

                valore_den = sheet[f'{colonna_den}{row}'].value

                if valore_den == '999':
                    valore_percentuale = 'null'
                    valore_percentuale_pesata = 'null'
                else:
                    try:
                        valore_percentuale = sheet[f'{colonna_percentuale}{row}'].value
                        valore_percentuale_pesata = float(valore_percentuale) * indicatore.peso / 100
                    except (ValueError, TypeError):
                        valore_percentuale = 'null'
                        valore_percentuale_pesata = 'null'

                sheet_uscita[f'{get_column_letter(Colonne.PERCENTUALE_PESATA)}{riga_uscita}'].value = valore_percentuale_pesata
                sheet_uscita[f"{get_column_letter(Colonne.ID)}{riga_uscita}"].value = id_riga
                sheet_uscita[f"{get_column_letter(Colonne.DATA)}{riga_uscita}"].value = data_riga
                sheet_uscita[f"{get_column_letter(Colonne.DATA_INVIO)}{riga_uscita}"].value = data_invio_riga
                sheet_uscita[f"{get_column_letter(Colonne.SETTING)}{riga_uscita}"].value = servizio_riga
                sheet_uscita[f'{get_column_letter(Colonne.PROFESSIONE)}{riga_uscita}'].value = professione_riga
                sheet_uscita[f"{get_column_letter(Colonne.ZONA)}{riga_uscita}"].value = zona_presidio_riga
                sheet_uscita[f"{get_column_letter(Colonne.REQUISITO)}{riga_uscita}"].value = requisito
                sheet_uscita[f"{get_column_letter(Colonne.SOC)}{riga_uscita}"].value = soc_riga
                sheet_uscita[f'{get_column_letter(Colonne.SOS)}{riga_uscita}'].value = sos_riga
                sheet_uscita[f'{get_column_letter(Colonne.TIPOLOGIA_PRESIDIO)}{riga_uscita}'].value = tipologia_presidio_riga
                sheet_uscita[f'{get_column_letter(Colonne.SEDE)}{riga_uscita}'].value = sede_presidio_riga

                sheet_uscita[f'{get_column_letter(Colonne.INDICATORE)}{riga_uscita}'].value = indicatore.stringa_uscita
                sheet_uscita[f"{get_column_letter(Colonne.NUMERATORE)}{riga_uscita}"].value = sheet[f'{colonna_num}{row}'].value
                sheet_uscita[f'{get_column_letter(Colonne.DENOMINATORE)}{riga_uscita}'].value = valore_den
                sheet_uscita[f'{get_column_letter(Colonne.PERCENTUALE)}{riga_uscita}'].value = valore_percentuale
                sheet_uscita[f'{get_column_letter(Colonne.PESO)}{riga_uscita}'].value = indicatore.peso
                riga_uscita += 1

    # Aggiusta la larghezza delle colonne
    for col in sheet_uscita.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        sheet_uscita.column_dimensions[column].width = adjusted_width

    wb_uscita.save(str(path_foglio_uscita))

    if is_csv_file:
        os.remove(_path_foglio_ingresso)


class Widget(QWidget):
    def __init__(self):
        super().__init__()

        layout = QVBoxLayout(self)

        hlayout = QHBoxLayout()

        self.__label_path_foglio_ingresso = QLabel("Nessun foglio di ingresso selezionato")
        self.__bottone_cambio_path_foglio_ingresso = QPushButton("...")
        hlayout.addWidget(self.__label_path_foglio_ingresso)
        hlayout.addWidget(self.__bottone_cambio_path_foglio_ingresso)

        self.__bottone_cambio_path_foglio_ingresso.clicked.connect(self.__selezionePathIngresso)

        layout.addLayout(hlayout)

        hlayout = QHBoxLayout()

        self.__label_path_foglio_uscita = QLabel("Nessun foglio di uscita selezionato")
        self.__bottone_cambio_path_foglio_uscita = QPushButton("...")

        self.__bottone_cambio_path_foglio_uscita.clicked.connect(self.__selezionePathUscita)

        hlayout.addWidget(self.__label_path_foglio_uscita)
        hlayout.addWidget(self.__bottone_cambio_path_foglio_uscita)

        self.__path_foglio_ingresso = None
        self.__path_foglio_uscita = None

        layout.addLayout(hlayout)

        self.__bottone_converti = QPushButton("Converti")
        self.__bottone_converti.setEnabled(False)
        self.__bottone_converti.clicked.connect(self.__converti)
        layout.addWidget(self.__bottone_converti)

    def __converti(self):
        converti(self.__path_foglio_ingresso, self.__path_foglio_uscita)

        QMessageBox.information(self,
                                "Conversione completata",
                                f"La conversione è stata completata nel file\n{self.__path_foglio_uscita}")

    def __updateStatoBottoneConverti(self):
        self.__bottone_converti.setEnabled(self.__path_foglio_uscita is not None and self.__path_foglio_ingresso is not None)

    def __selezionePathIngresso(self):
        fnames, _ = QFileDialog.getOpenFileNames(self, "Seleziona il file di ingresso",
                                                 filter="CSV (*.csv);;Excel (*.xlsx)")

        if (fname := next(iter(fnames), None)) is not None:
            self.__path_foglio_ingresso = Path(fname)
            self.__label_path_foglio_ingresso.setText(str(self.__path_foglio_ingresso))

        self.__updateStatoBottoneConverti()

    def __selezionePathUscita(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Seleziona dove salvare il file", filter="Excel (*.xlsx)")

        if fname:
            self.__path_foglio_uscita = Path(fname)
            self.__label_path_foglio_uscita.setText(str(self.__path_foglio_uscita))

        self.__updateStatoBottoneConverti()


def main():
    app = QApplication([])
    w = Widget()

    w.show()
    app.exec()


if __name__ == "__main__":
    main()
