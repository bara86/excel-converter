from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, NamedTuple
from enum import IntEnum, auto

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from PySide6.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFileDialog

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


class Colonne(IntEnum):
    ID = 1
    DATA = auto()
    PROFESSIONE = auto()
    SOC = auto()
    SOS = auto()
    ZONA = auto()
    TIPOLOGIA_PRESIDIO = auto()
    SEDE = auto()
    REQUISITO = auto()
    INDICATORE = auto()
    NUMERATORE = auto()
    DENOMINATORE = auto()


class Indicatore(NamedTuple):
    stringa_uscita: str
    nome_indicatore: str


mappa_requisito_indicatori: dict[str, list[Indicatore]] = {
    'DISPOSITIVI MEDICI': [Indicatore('MANUTENZIONE TSLB', 'manutenzione_tslb')],
    'NEOINSERITO / NEOASSUNTO': [Indicatore('PIANO INSERIMENTO NEOASSUNTO', 'neoassunto'),
                                 Indicatore('PIANO INSERIMENTO NEOINSERITO', 'neoinserito')]
}


def cercaRigaColonna(sheet: Worksheet, nome_colonna: str) -> str | None:
    for col in range(1, sheet.max_column + 1):
        col_name = get_column_letter(col)
        if sheet[f"{col_name}1"].value.strip() == nome_colonna:
            return col_name


def scriviTitoliUscita(sheet: Worksheet) -> None:
    for colonna in Colonne:
        sheet[f'{get_column_letter(colonna)}1'].value = colonna.name.replace('_', ' ')


def converti(path_foglio_ingresso: Path, path_foglio_uscita: Path) -> None:
    wb: Workbook = load_workbook(str(path_foglio_ingresso))

    sheet: Worksheet = wb.worksheets[0]
    indice_id = cercaRigaColonna(sheet, 'id')
    indice_zona_presidio = cercaRigaColonna(sheet, 'params.zona_presidio')
    indice_professione = cercaRigaColonna(sheet, 'params.Professione')
    indice_data = cercaRigaColonna(sheet, 'data')
    indice_sos = cercaRigaColonna(sheet, 'params.SOS')
    indice_soc = cercaRigaColonna(sheet, 'params.SOC')
    indice_tipologia_presidio = cercaRigaColonna(sheet, 'params.presidio')
    indice_sede_presidio = cercaRigaColonna(sheet, 'params.sede_presidio')

    wb_uscita = Workbook()
    sheet_uscita = wb_uscita.worksheets[0]

    scriviTitoliUscita(sheet_uscita)

    riga_uscita = 2
    for row in range(2, sheet.max_row + 1):

        for (requisito, indicatori) in mappa_requisito_indicatori.items():
            id_riga = sheet[f"{indice_id}{row}"].value
            zona_presidio_riga = sheet[f"{indice_zona_presidio}{row}"].value
            data_riga = sheet[f"{indice_data}{row}"].value.strftime("%d/%m/%Y")
            professione_riga = sheet[f'{indice_professione}{row}'].value
            soc_riga = sheet[f'{indice_soc}{row}'].value
            sos_riga = sheet[f'{indice_sos}{row}'].value
            tipologia_presidio_riga = sheet[f'{indice_tipologia_presidio}{row}'].value
            sede_presidio_riga = sheet[f'{indice_sede_presidio}{row}'].value

            for contatore_uscita, indicatore in enumerate(indicatori):
                sheet_uscita[f"{get_column_letter(Colonne.ID)}{riga_uscita}"].value = id_riga
                sheet_uscita[f"{get_column_letter(Colonne.DATA)}{riga_uscita}"].value = data_riga
                sheet_uscita[f'{get_column_letter(Colonne.PROFESSIONE)}{riga_uscita}'].value = professione_riga
                sheet_uscita[f"{get_column_letter(Colonne.ZONA)}{riga_uscita}"].value = zona_presidio_riga
                sheet_uscita[f"{get_column_letter(Colonne.REQUISITO)}{riga_uscita}"].value = requisito
                sheet_uscita[f"{get_column_letter(Colonne.SOC)}{riga_uscita}"].value = soc_riga
                sheet_uscita[f'{get_column_letter(Colonne.SOS)}{riga_uscita}'].value = sos_riga
                sheet_uscita[f'{get_column_letter(Colonne.TIPOLOGIA_PRESIDIO)}{riga_uscita}'].value = tipologia_presidio_riga
                sheet_uscita[f'{get_column_letter(Colonne.SEDE)}{riga_uscita}'].value = sede_presidio_riga

                colonna_num = cercaRigaColonna(sheet, f"params.num_{indicatore.nome_indicatore}")
                colonna_den = cercaRigaColonna(sheet, f'params.den_{indicatore.nome_indicatore}')
                sheet_uscita[f'{get_column_letter(Colonne.INDICATORE)}{riga_uscita}'].value = indicatore.stringa_uscita
                sheet_uscita[f"{get_column_letter(Colonne.NUMERATORE)}{riga_uscita}"].value = sheet[f'{colonna_num}{row}'].value
                sheet_uscita[f'{get_column_letter(Colonne.DENOMINATORE)}{riga_uscita}'].value = sheet[f'{colonna_den}{row}'].value
                riga_uscita += 1

    # Aggiusta la larghezza delle colonne
    for col in sheet_uscita.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        sheet_uscita.column_dimensions[column].width = adjusted_width

    wb_uscita.save(str(path_foglio_uscita))


class Widget(QWidget):
    def __init__(self):
        super().__init__()

        layout = QVBoxLayout(self)

        hlayout = QHBoxLayout()

        self.__label_path_foglio_ingresso = QLabel("Nessun foglio selezionato")
        self.__bottone_cambio_path_foglio_ingresso = QPushButton("...")
        hlayout.addWidget(self.__label_path_foglio_ingresso)
        hlayout.addWidget(self.__bottone_cambio_path_foglio_ingresso)

        self.__bottone_cambio_path_foglio_ingresso.clicked.connect(self.__selezionePathIngresso)

        layout.addLayout(hlayout)

        hlayout = QHBoxLayout()

        self.__label_path_foglio_uscita = QLabel("Nessun foglio di uscita selezionato")
        self.__bottone_cambio_path_foglio_uscita = QPushButton("...")

        self.__bottone_cambio_path_foglio_uscita.clicked.connect(self.__selezionePathUscita)

        hlayout.addWidget(self.__label_path_foglio_uscita)
        hlayout.addWidget(self.__bottone_cambio_path_foglio_uscita)

        self.__path_foglio_ingresso = Path(r"C:\Users\Bara\Downloads\FILE CONVERT.xlsx")
        self.__path_foglio_uscita = Path(r"C:\Users\Bara\Downloads\file_convertito.xlsx")

        layout.addLayout(hlayout)

        self.__bottone_converti = QPushButton("Converti")
        self.__bottone_converti.setEnabled(True)
        self.__bottone_converti.clicked.connect(self.__converti)
        layout.addWidget(self.__bottone_converti)

    def __converti(self):
        converti(self.__path_foglio_ingresso, self.__path_foglio_uscita)

    def __updateStatoBottoneConverti(self):
        self.__bottone_converti.setEnabled(not (self.__label_path_foglio_uscita is not None and self.__label_path_foglio_ingresso is not None))

    def __selezionePathIngresso(self):
        fnames, _ = QFileDialog.getOpenFileNames(self, "Seleziona il file di ingresso", filter="Excel (*.xls *.xlsx)")

        if fname := fnames[0]:
            self.__path_foglio_ingresso = Path(fname)
            self.__label_path_foglio_ingresso.setText(str(self.__path_foglio_ingresso))

        self.__updateStatoBottoneConverti()

    def __selezionePathUscita(self):
        fname, _ = QFileDialog.getSaveFileName(self, "Seleziona dove salvare il file", filter="Excel (*.xls *.xlsx)")

        if fname:
            self.__path_foglio_uscita = Path(fname)
            self.__label_path_foglio_uscita.setText(str(self.__path_foglio_uscita))

        self.__updateStatoBottoneConverti()


def main():
    app = QApplication([])
    w = Widget()

    w.show()
    app.exec()


if __name__ == "__main__":
    main()
